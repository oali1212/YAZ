from openpyxl import load_workbook, workbook, Workbook
from openpyxl.utils import *
from openpyxl.styles import *
import os
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from configparser import ConfigParser
from PyQt5.QtGui import QFont
import math
from datetime import datetime
import pandas as pd
import sys
class YAZ:


    def __init__(self): 


        self.main_sheet_titles = ["Date", "Time", "Unique ID", "Transaction Type", "Transaction Description","Amount (EGP)",    "Logged by:" , "Notes"]

        self.excel_file_path = 'Customers Bills//main_sheet.xlsx'
        self.excel_file_path = self.get_relink(self.excel_file_path)
    def create_main_sheet(self):
        # Define the path to save the workbook

        
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Set the title of the worksheet
        ws.title = "Main Sheet"
        for col_num, title in enumerate(self.main_sheet_titles, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.font = Font(bold=True)        
    
   
        # Save the workbook
        try:
            wb.save(self.excel_file_path)
            return True
        except Exception as e:
            return str(e)

    def resize_columns_to_fit_content(self,ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Adjusting width a little bit for padding
            ws.column_dimensions[column].width = adjusted_width
            # Center align all cells in the new row
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def append_row_to_main_sheet(self, row_data):
        # Define the path to save the workbook
        directory_path = 'Customers Bills'
        directory_path = self.get_relink(directory_path)

        
        # Check if the row has the correct length
        if len(row_data) != len(self.main_sheet_titles):
            return False
        
        # Check if the workbook exists, if not create it
        if not os.path.exists(self.excel_file_path):
            self.create_main_sheet()
        
        # Load the workbook and select the active worksheet
        wb = load_workbook(self.excel_file_path)
        ws = wb.active
        self.resize_columns_to_fit_content(ws)   
        # Append the row data
        ws.append(row_data)
        
        # Save the workbook
        try:
            wb.save(self.excel_file_path)
            return True
        except Exception as e: 
            return str(e)
        
        return True

    def delete_row_by_id(self, value):
        # Define the path to save the workbook
        directory_path = 'Customers Bills'
        directory_path = self.get_relink(directory_path)

        
        # Check if the workbook exists
        if not os.path.exists(self.excel_file_path):
            print(f"Workbook does not exist at path: {self.excel_file_path}")
            return False
        
        print(f"Loading workbook from path: {self.excel_file_path}")
        
        # Load the workbook and select the active worksheet
        wb = load_workbook(self.excel_file_path)
        ws = wb.active
        
        print(f"Searching for value '{value}' in worksheet...")
        
        # Iterate over the rows to find the value in the second column
        row_to_delete = None
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=False):
            if row[2].value == value:
                row_to_delete = row[0].row
                print(f"Found matching value '{value}' in row {row_to_delete}")
                break
        
        # If a matching row is found, delete it
        if row_to_delete:
            ws.delete_rows(row_to_delete)
            print(f"Deleted row {row_to_delete} containing value '{value}'")
            try:
                wb.save(self.excel_file_path)
                print(f"Workbook saved successfully at path: {self.excel_file_path}")
                return True
            except Exception as e:
                error_message = str(e)
                print(f"Failed to save workbook: {error_message}")
                return error_message
        else:
            print(f"Value '{value}' not found in worksheet.")
            return False
    def create_price_table(self, ini_file, section):
        config = ConfigParser()
        config.read(ini_file)

        if section not in config:
            ##print(f"Section {section} not found in the INI file.")
            return None
        
        # Assuming services are stored in a "key = value" format under section
        services = config[section]
        
        # Create Table
        tableWidget = QTableWidget()
        
        # Set table dimensions
        tableWidget.setRowCount(len(services))
        tableWidget.setColumnCount(5)
        tableWidget.setHorizontalHeaderLabels(['Service', 'Price (EGP)', 'Price -30%', 'Price -40%', "Remove"])
        tableWidget.horizontalHeader().setStyleSheet("font-weight: bold; font-size:25px; background-color: lightgrey;")
        # Hide the vertical header (numbers column)
        tableWidget.verticalHeader().setVisible(False)

        font = QFont()
        font.setPointSize(25)
        for i, (service, price) in enumerate(services.items()):
            price = float(price)
            price_discounted_30 = math.ceil( price * 0.7 )
            price_discounted_40 = math.ceil( price * 0.6 )
            
            service = service.replace("__", " + ")
            service = service.replace("_", " ")
            service = service.upper()

            self.remove = QPushButton("🗑")
            self.remove.setStyleSheet("font-size:25px;")


            tableWidget.setItem(i, 0, QTableWidgetItem(service))
            price = str(price)
            price_discounted_30 = str(price_discounted_30)
            price_discounted_40 = str(price_discounted_40)
            tableWidget.setItem(i, 1, QTableWidgetItem(price))
            tableWidget.setItem(i, 2, QTableWidgetItem(price_discounted_30))
            tableWidget.setItem(i, 3, QTableWidgetItem(price_discounted_40))
            tableWidget.setCellWidget(i,4,self.remove)


        tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        tableWidget.setStyleSheet("background-color: white;")
        # Enable resizing of columns by mouse
        tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        # Set the horizontal header resize mode to Stretch
        tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # Set the vertical header resize mode to Stretch
        tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

    


        tableWidget.setFixedWidth(1100)
        self.disable_modification(tableWidget)
        return tableWidget
    
    def create_customers_table(self, ini_file):
        tableWidget = QTableWidget()
        tableWidget.setStyleSheet("background-color: white;")

        # Load data from ini file
        config = ConfigParser()
        config.read(ini_file)

        # Set table properties
        tableWidget.setColumnCount(7)
        tableWidget.horizontalHeader().setFont(QFont('Arial', 12, QFont.Bold))
        tableWidget.horizontalHeader().setStyleSheet("background-color: lightgrey; font-weight: bold; font-size: 12px;")
        tableWidget.verticalHeader().setVisible(False)

        # Iterate over each section (customer) in the INI file
        for section in config.sections():
            # Retrieve all options for the current section
            options = [section] + [config[section][option] for option in config[section]]
            # Append options to the table as a new row
            row_position = tableWidget.rowCount()
            tableWidget.insertRow(row_position)
            for j, option in enumerate(options):
                item = QTableWidgetItem(option)
                tableWidget.setItem(row_position, j, item)

        # Set table dimensions and properties
        tableWidget.setRowCount(tableWidget.rowCount())  # Ensure proper row count
        tableWidget.setHorizontalHeaderLabels(["Name", "#", "Last Date Vistied", "Phone Number", "Total Paid", "Email", "Heard about us from:"])
        tableWidget.setFixedWidth(1100)
        tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

        return tableWidget

    def update_ini_from_table(self,tableWidget, ini_file):
        # Initialize ConfigParser
        config = ConfigParser()
        
        # Iterate over each row in the table
        for row in range(tableWidget.rowCount()):
            # Get the section name (first column)
            if tableWidget.item(row,0):
                section = tableWidget.item(row, 0).text()
                if section != '' and section != ' ':
                    config[section] = {}

                    # Populate the rest of the options
                    config[section]['Index'] = tableWidget.item(row, 1).text()
                    config[section]['date_registered'] = tableWidget.item(row, 2).text()
                    config[section]['phone_number'] = tableWidget.item(row, 3).text()
                    config[section]['total_paid'] = tableWidget.item(row, 4).text()
                    config[section]['email'] = tableWidget.item(row, 5).text()
                    config[section]['heard_about_us'] = tableWidget.item(row, 6).text()
                
                    ##print(config.sections())
            # Write the updated data back to the INI file
            #print(ini_file)
            with open(ini_file, 'w') as configfile:
                
                config.write(configfile)
                ##print(f"INI file '{ini_file}' updated successfully.")

    def disable_modification(self, table):
        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                item = table.item(row, col)
                if item:
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)

    def add_service(self, ini_file, section, name, price):

        config = ConfigParser()
        config.read(ini_file)
        
        if not config.has_section(section):
            config.add_section(section)
        
        if name[-1] == "_":
            name = name[:-2]

        config.set(section, name, str(price))
        
        with open(ini_file, 'w') as config_file:
            config.write(config_file)

    def delete_service(self, path, section, index):
        index = int(index)
        config = ConfigParser()
        config.read(path)
 
        if section not in config:
            raise ValueError(f"Section '{section}' not found in the INI file.")

        options = list(config[section])
        if index < 1 or index > len(options):
            raise ValueError("Invalid index.")

        option_to_delete = options[index - 1]  # Adjust index to 0-based index
        del config[section][option_to_delete]

        with open(path, 'w') as configfile:
            config.write(configfile)
        return True
    def save_services(self, path, section, prices_list: list):
        config = ConfigParser()
        config.read(path)

        if section not in config:
            raise ValueError(f"Section '{section}' not found in the INI file.")
        
        options = list(config[section])
        
        if len(options) != len(prices_list):
            raise ValueError("Length of prices list does not match the number of options in the section.")
        
        for index, price in enumerate(prices_list):
            option_to_edit = options[index]  # Zero-based index
            config[section][option_to_edit] = str(price)
        
        # Write changes back to the file
        with open(self.excel_file_path, 'w') as configfile:
            config.write(configfile)

    def save_table_to_excel(self, table: QTableWidget, name: str):
        # Convert QTableWidget to pandas DataFrame
        data = []
        for row in range(table.rowCount()):
            row_data = []
            for column in range(table.columnCount()):
                item = table.item(row, column)
                row_data.append(item.text() if item is not None else '')
            data.append(row_data)
        
        headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
        df = pd.DataFrame(data, columns=headers)
        
        # Check if folder for Customers Bills exists, if not create it
        customers_bills_folder = "Customers Bills"
        customers_bills_folder = self.get_relink(customers_bills_folder)
        if not os.path.exists(customers_bills_folder):
            os.makedirs(customers_bills_folder)
        
        # Check if folder for the customer exists, if not create it
        customer_folder = os.path.join(customers_bills_folder, name)
        if not os.path.exists(customer_folder):
            os.makedirs(customer_folder)
        
        # Save with a timestamped name
        timestamp = datetime.now().strftime("%d_%m_%y_%H_%M_%S")
        file_name = f"{name}_{timestamp}.xlsx"
        user_file_path = os.path.join(customers_bills_folder,name, file_name)
        df.to_excel(user_file_path, index=False)
        #print(f"File saved as {self.excel_file_path}")
        # Open Microsoft Excel 

    def get_table_from_excel(self, path):
        if not os.path.exists(path):
            return None

        try:
            wb = load_workbook(path)
            sheet = wb.active  # Get the active sheet (first sheet by default)

            rows = sheet.iter_rows(values_only=True)

            num_rows = sheet.max_row
            num_cols = sheet.max_column

            table_widget = QTableWidget()
            table_widget.setRowCount(num_rows)
            table_widget.setColumnCount(num_cols)
            table_widget.horizontalHeader().setVisible(False)

            for row_index, row in enumerate(rows):
                for col_index, value in enumerate(row):
                    item = QTableWidgetItem(str(value))
                    table_widget.setItem(row_index, col_index, item)
                    item = table_widget.item(row_index,col_index)
                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)

            table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            table_widget.setStyleSheet("background-color: white;")
            return table_widget
        except Exception as e:
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Critical)
            msg.setText(f"Error loading Excel file: {e}")
            msg.setWindowTitle("Error")
            msg.exec_()
            return None
        


    def get_relink(self,link):
        # determine if application is a script file or frozen exe
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
        elif __file__:
            application_path = os.path.dirname(__file__)

        config_path = os.path.join(application_path, link)

        return config_path



# yaz.create_main_sheet()
# yaz.append_row_to_main_sheet([1,2223232,3,4,5,6,7,8])
# yaz.append_row_to_main_sheet([1,2222,3,4,5,6,7,8])
# yaz.delete_row_by_id(2222)